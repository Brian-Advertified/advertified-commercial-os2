"""Upload certified supplier workbooks through the normal inventory API."""

from __future__ import annotations

import argparse
import hashlib
import json
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from inventory_certified_upload_client import CertifiedInventoryApi, deterministic_uuid

ROOT = Path(__file__).resolve().parents[1]
CORPUS = ROOT / "artifacts" / "inventory-corpus"
UPLOAD_ROOT = CORPUS / "certified-upload"
UPLOAD_MARKER = UPLOAD_ROOT / "UPLOAD_FILES_VERIFIED.json"
PLAN_PATH = UPLOAD_ROOT / "upload-plan.json"
CONTRACT_PATH = UPLOAD_ROOT / "api-contract.json"
STATE_PATH = UPLOAD_ROOT / "database-upload-state.json"
RESULT_ROOT = UPLOAD_ROOT / "database-results"
CANARY_MARKER = UPLOAD_ROOT / "DMS_DATABASE_UPLOAD_PASSED.json"
COMPLETE_MARKER = UPLOAD_ROOT / "DATABASE_UPLOAD_RECONCILED.json"
DMS_SUPPLIER = "DStv Media Sales"
TENANT_ID = "10000000-0000-0000-0000-000000000020"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--all", action="store_true")
    parser.add_argument("--api", default="http://127.0.0.1:5197")
    parser.add_argument("--origin", default="http://localhost:3017")
    parser.add_argument("--tenant", default=TENANT_ID)
    args = parser.parse_args()
    require(UPLOAD_MARKER, "uploadFilesVerified")
    if args.all:
        require(CANARY_MARKER, "dmsDatabaseUploadPassed")
    plan = read_json(PLAN_PATH)
    contract = read_json(CONTRACT_PATH)
    state = read_json(STATE_PATH) if STATE_PATH.exists() else new_state(plan)
    client = CertifiedInventoryApi(
        args.api, args.origin, args.tenant, contract
    )
    client.authenticate()
    workbooks = list(plan.get("workbooks") or [])
    selected = workbooks if args.all else [
        item for item in workbooks
        if item.get("supplierName") == DMS_SUPPLIER
    ]
    if not selected:
        raise RuntimeError("The certified DMS supplier workbook is missing.")
    RESULT_ROOT.mkdir(parents=True, exist_ok=True)
    for entry in selected:
        process_workbook(client, state, entry)
        write_json_atomic(STATE_PATH, state)
    if args.all:
        reconciliation = reconcile_all(plan, state)
        write_json(UPLOAD_ROOT / "database-upload-reconciliation.json", reconciliation)
        if not reconciliation["passed"]:
            COMPLETE_MARKER.unlink(missing_ok=True)
            raise RuntimeError("The certified database upload did not reconcile.")
        write_json(COMPLETE_MARKER, {
            "databaseUploadReconciled": True,
            "sourceCount": reconciliation["sourceCount"],
            "supplierCount": reconciliation["supplierCount"],
            "productCount": reconciliation["expectedProductCount"],
            "uploadedCandidateCount": reconciliation["uploadedCandidateCount"],
            "importCount": reconciliation["importCount"],
            "allAtHumanReview": reconciliation["allAtHumanReview"],
            "nothingPublished": reconciliation["nothingPublished"],
        })
        print(json.dumps(reconciliation, indent=2))
    else:
        dms = next(
            item for item in state["workbooks"]
            if item["supplierName"] == DMS_SUPPLIER
        )
        if not dms.get("reconciled"):
            raise RuntimeError("The DMS database upload canary did not reconcile.")
        write_json(CANARY_MARKER, {
            "dmsDatabaseUploadPassed": True,
            "supplierName": DMS_SUPPLIER,
            "importId": dms["importId"],
            "workbookSha256": dms["workbookSha256"],
            "expectedProductCount": dms["expectedProductCount"],
            "uploadedCandidateCount": dms["uploadedCandidateCount"],
            "status": dms["status"],
        })
        print(json.dumps(dms, indent=2))
    return 0


def process_workbook(
    client: CertifiedInventoryApi,
    state: dict[str, Any],
    entry: dict[str, Any],
) -> None:
    workbook_path = ROOT / str(entry["workbookPath"])
    expected_codes = workbook_product_codes(workbook_path)
    expected_count = int(entry["productCount"])
    if len(expected_codes) != expected_count:
        raise RuntimeError(
            f"Workbook product-code count mismatch for {entry['supplierName']}."
        )
    workbook_hash = file_hash(workbook_path)
    if workbook_hash != entry["workbookSha256"]:
        raise RuntimeError(
            f"Workbook hash changed for {entry['supplierName']}."
        )
    state_entry = next((
        item for item in state["workbooks"]
        if item["workbookSha256"] == workbook_hash
    ), None)
    if state_entry and state_entry.get("reconciled"):
        verify_existing(client, state_entry, expected_codes)
        return
    if state_entry is None:
        state_entry = {
            "supplierName": entry["supplierName"],
            "workbookPath": entry["workbookPath"],
            "workbookSha256": workbook_hash,
            "expectedProductCount": expected_count,
            "sourceHashes": entry.get("sourceHashes") or [],
            "status": "PENDING",
            "reconciled": False,
        }
        state["workbooks"].append(state_entry)
        write_json_atomic(STATE_PATH, state)

    supplier_id = client.ensure_supplier(str(entry["supplierName"]))
    state_entry["supplierId"] = supplier_id
    import_id = state_entry.get("importId")
    if import_id:
        current = client.read_complete_import(str(import_id))
    else:
        created = client.create_import(
            workbook_path,
            str(entry["supplierName"]),
            supplier_id,
            deterministic_uuid("upload:" + workbook_hash),
        )
        import_id = str(created.get("importId") or created.get("id"))
        state_entry["importId"] = import_id
        state_entry["createdAtUtc"] = datetime.now(UTC).isoformat()
        state_entry["status"] = str(created.get("status") or "UPLOADED")
        write_json_atomic(STATE_PATH, state)
        current = wait_for_scan_or_review(client, import_id)

    status = str(current.get("status") or "")
    if status == "UPLOADED":
        version = int(current.get("version") or 0)
        client.execute_import(
            import_id,
            version,
            deterministic_uuid("execute:" + workbook_hash),
        )
        current = client.wait_for_import(import_id)
    elif status in {"EXTRACTING", "PENDING", "RUNNING"}:
        current = client.wait_for_import(import_id)
    reconcile_import(
        state_entry,
        current,
        expected_codes,
        workbook_hash,
    )
    write_json(
        RESULT_ROOT / f"{workbook_hash}.json",
        {
            "workbook": entry,
            "state": state_entry,
            "import": current,
        },
    )


def wait_for_scan_or_review(
    client: CertifiedInventoryApi,
    import_id: str,
    timeout_seconds: int = 300,
) -> dict[str, Any]:
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        current = client.read_complete_import(import_id)
        status = str(current.get("status") or "")
        scan = str(current.get("scanStatus") or "")
        if status == "REVIEW_REQUIRED":
            return current
        if status == "UPLOADED" and scan in {"CLEAN", "COMPLETED", "PASSED", ""}:
            return current
        if status in {"FAILED", "FAILED_TERMINAL", "CANCELLED"}:
            return current
        time.sleep(2)
    raise TimeoutError(f"Malware scan did not complete for {import_id}.")


def reconcile_import(
    state_entry: dict[str, Any],
    current: dict[str, Any],
    expected_codes: set[str],
    workbook_hash: str,
) -> None:
    candidates = current.get("candidates") or []
    observed_codes = []
    for candidate in candidates:
        values = (
            candidate.get("canonicalValues")
            or candidate.get("proposedValues")
            or candidate.get("values")
            or {}
        )
        code = str(values.get("productCode") or "")
        if code:
            observed_codes.append(code)
    failures = []
    if str(current.get("status") or "") != "REVIEW_REQUIRED":
        failures.append("IMPORT_NOT_AT_HUMAN_REVIEW")
    if current.get("failureCode"):
        failures.append("IMPORT_FAILURE_PRESENT")
    if len(candidates) != len(expected_codes):
        failures.append("CANDIDATE_COUNT_MISMATCH")
    if set(observed_codes) != expected_codes:
        failures.append("PRODUCT_CODE_SET_MISMATCH")
    if len(observed_codes) != len(set(observed_codes)):
        failures.append("DUPLICATE_PRODUCT_CODES")
    state_entry.update({
        "workbookSha256": workbook_hash,
        "status": current.get("status"),
        "failureCode": current.get("failureCode"),
        "uploadedCandidateCount": len(candidates),
        "observedProductCodeCount": len(observed_codes),
        "reconciliationFailures": failures,
        "reconciled": not failures,
        "reconciledAtUtc": datetime.now(UTC).isoformat(),
        "nothingPublished": all(
            str(candidate.get("status") or "") != "PUBLISHED"
            for candidate in candidates
        ),
    })
    if failures:
        raise RuntimeError(
            f"Certified import {state_entry.get('importId')} failed reconciliation: "
            + ", ".join(failures)
        )


def verify_existing(
    client: CertifiedInventoryApi,
    state_entry: dict[str, Any],
    expected_codes: set[str],
) -> None:
    current = client.read_complete_import(str(state_entry["importId"]))
    reconcile_import(
        state_entry,
        current,
        expected_codes,
        str(state_entry["workbookSha256"]),
    )


def reconcile_all(plan: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    planned = {str(item["workbookSha256"]): item for item in plan.get("workbooks") or []}
    completed = {
        str(item["workbookSha256"]): item
        for item in state.get("workbooks") or []
        if item.get("reconciled")
    }
    missing = sorted(set(planned) - set(completed))
    extra = sorted(set(completed) - set(planned))
    expected_count = sum(int(item["productCount"]) for item in planned.values())
    observed_count = sum(int(item.get("uploadedCandidateCount") or 0) for item in completed.values())
    statuses = {str(item.get("status") or "") for item in completed.values()}
    failures = []
    if missing:
        failures.append("WORKBOOK_IMPORTS_MISSING")
    if extra:
        failures.append("UNPLANNED_WORKBOOK_IMPORTS_PRESENT")
    if expected_count != observed_count:
        failures.append("TOTAL_CANDIDATE_COUNT_MISMATCH")
    if statuses != {"REVIEW_REQUIRED"}:
        failures.append("NOT_ALL_IMPORTS_AT_HUMAN_REVIEW")
    if not all(item.get("nothingPublished") for item in completed.values()):
        failures.append("CANDIDATES_PUBLISHED_BEFORE_RECONCILIATION")
    return {
        "passed": not failures,
        "failures": failures,
        "sourceCount": int(plan["sourceCount"]),
        "supplierCount": int(plan["supplierCount"]),
        "importCount": len(completed),
        "expectedImportCount": len(planned),
        "expectedProductCount": expected_count,
        "uploadedCandidateCount": observed_count,
        "allAtHumanReview": statuses == {"REVIEW_REQUIRED"},
        "nothingPublished": all(item.get("nothingPublished") for item in completed.values()),
        "missingWorkbookHashes": missing,
        "extraWorkbookHashes": extra,
    }


def workbook_product_codes(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Certified Inventory"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    index = headers.index("product_code")
    values = {
        str(row[index]) for row in rows
        if index < len(row) and row[index]
    }
    workbook.close()
    return values


def new_state(plan: dict[str, Any]) -> dict[str, Any]:
    return {
        "schemaVersion": "advertified.certified-inventory-database-upload.v1",
        "createdAtUtc": datetime.now(UTC).isoformat(),
        "uploadPlanSha256": file_hash(PLAN_PATH),
        "expectedSupplierCount": plan.get("supplierCount"),
        "expectedProductCount": plan.get("canonicalProductCount"),
        "workbooks": [],
    }


def require(path: Path, field: str) -> dict[str, Any]:
    value = read_json(path)
    if not value.get(field):
        raise RuntimeError(f"Required certificate {path.name} is invalid.")
    return value


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.resolve(strict=True).read_bytes()).hexdigest()


def read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.resolve(strict=True).read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"Expected object in {path}")
    return value


def write_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def write_json_atomic(path: Path, value: dict[str, Any]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    write_json(temporary, value)
    temporary.replace(path)


if __name__ == "__main__":
    raise SystemExit(main())

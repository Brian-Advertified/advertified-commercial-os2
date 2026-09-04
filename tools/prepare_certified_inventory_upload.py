"""Create and verify supplier upload workbooks from certified canonical inventory."""

from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
CORPUS = ROOT / "artifacts" / "inventory-corpus"
CANONICAL_ROOT = CORPUS / "certified-canonical-inventory"
CANONICAL_PATH = CANONICAL_ROOT / "canonical-products.json"
CANONICAL_MARKER = CANONICAL_ROOT / "ALL_CANONICAL_PRODUCTS_ASSEMBLED.json"
OUTPUT_ROOT = CORPUS / "certified-upload"
WORKBOOK_ROOT = OUTPUT_ROOT / "supplier-workbooks"
PLAN_PATH = OUTPUT_ROOT / "upload-plan.json"
READY_MARKER = OUTPUT_ROOT / "UPLOAD_FILES_VERIFIED.json"

COLUMNS = [
    "supplier_name", "product_code", "name", "channel", "product_type",
    "geography", "description", "rate_type", "currency", "rate_raw",
    "rate_amount_minor", "availability", "valid_from", "valid_to",
    "source_hash", "source_file_name", "source_locators",
    "source_row_ids", "ambiguity_codes", "review_notes",
    "bedrock_packet_id", "bedrock_prompt_version", "bedrock_request_hash",
]


def main() -> int:
    marker = require(CANONICAL_MARKER, "allCanonicalProductsAssembled")
    canonical = read_json(CANONICAL_PATH)
    if file_hash(CANONICAL_PATH) != marker.get("canonicalInventorySha256"):
        raise RuntimeError("Canonical inventory hash does not match its certificate.")
    products = canonical.get("products") or []
    if not products:
        raise RuntimeError("Certified canonical inventory contains no products.")
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        grouped[str(product["supplierName"])].append(product)
    WORKBOOK_ROOT.mkdir(parents=True, exist_ok=True)
    for old in WORKBOOK_ROOT.glob("*.xlsx"):
        old.unlink()
    entries = []
    for supplier, supplier_products in sorted(grouped.items()):
        path = WORKBOOK_ROOT / f"{slug(supplier)}.certified-inventory.xlsx"
        write_workbook(path, supplier_products)
        read_back = read_workbook(path)
        expected = [workbook_row(product) for product in supplier_products]
        if read_back != expected:
            raise RuntimeError(f"Workbook read-back mismatch for {supplier}.")
        entries.append({
            "supplierName": supplier,
            "productCount": len(supplier_products),
            "sourceHashes": sorted({str(item["sourceHash"]) for item in supplier_products}),
            "workbookPath": str(path.relative_to(ROOT)),
            "workbookSha256": file_hash(path),
            "workbookBytes": path.stat().st_size,
        })
    plan = {
        "schemaVersion": "advertified.certified-inventory-upload-plan.v1",
        "generatedAtUtc": datetime.now(UTC).isoformat(),
        "sourceCount": int(canonical["sourceCount"]),
        "supplierCount": len(entries),
        "canonicalProductCount": len(products),
        "workbookProductCount": sum(item["productCount"] for item in entries),
        "canonicalInventorySha256": file_hash(CANONICAL_PATH),
        "bedrockProgrammeUsedOrReservedUsdMicros": canonical[
            "bedrockProgrammeUsedOrReservedUsdMicros"
        ],
        "bedrockProgrammeLimitUsdMicros": canonical[
            "bedrockProgrammeLimitUsdMicros"
        ],
        "workbooks": entries,
    }
    if plan["sourceCount"] != 43:
        raise RuntimeError("Upload plan does not cover all 43 physical sources.")
    if plan["canonicalProductCount"] != plan["workbookProductCount"]:
        raise RuntimeError("Upload workbook product count does not reconcile.")
    write_json(PLAN_PATH, plan)
    write_json(READY_MARKER, {
        "uploadFilesVerified": True,
        "sourceCount": plan["sourceCount"],
        "supplierCount": plan["supplierCount"],
        "productCount": plan["canonicalProductCount"],
        "uploadPlanSha256": file_hash(PLAN_PATH),
        "canonicalInventorySha256": plan["canonicalInventorySha256"],
    })
    print(json.dumps({
        "sourceCount": plan["sourceCount"],
        "supplierCount": plan["supplierCount"],
        "productCount": plan["canonicalProductCount"],
        "workbookCount": len(entries),
        "output": str(PLAN_PATH.relative_to(ROOT)),
    }, indent=2))
    return 0


def write_workbook(path: Path, products: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Certified Inventory"
    sheet.append(COLUMNS)
    for product in sorted(products, key=lambda item: (item["sourceHash"], item["name"], item["productCode"])):
        values = workbook_row(product)
        sheet.append([values[column] for column in COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:W{sheet.max_row}"
    widths = {
        "A": 28, "B": 24, "C": 45, "D": 18, "E": 26,
        "F": 35, "G": 70, "H": 20, "I": 12, "J": 20,
        "K": 20, "L": 18, "M": 15, "N": 15, "O": 68,
        "P": 45, "Q": 55, "R": 55, "S": 30, "T": 45,
        "U": 68, "V": 36, "W": 68,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    workbook.save(path)


def read_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Certified Inventory"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    result = []
    for values in rows:
        result.append({
            header: normalize_cell(value)
            for header, value in zip(headers, values, strict=True)
        })
    workbook.close()
    return result


def workbook_row(product: dict[str, Any]) -> dict[str, Any]:
    bedrock = product.get("bedrock") or {}
    return {
        "supplier_name": product["supplierName"],
        "product_code": product["productCode"],
        "name": product["name"],
        "channel": product["channel"],
        "product_type": product["productType"],
        "geography": product.get("geography"),
        "description": product.get("description"),
        "rate_type": product.get("rateType"),
        "currency": product.get("currency"),
        "rate_raw": product.get("rateRaw"),
        "rate_amount_minor": product.get("rateAmountMinor"),
        "availability": product.get("availability"),
        "valid_from": product.get("validFrom"),
        "valid_to": product.get("validTo"),
        "source_hash": product["sourceHash"],
        "source_file_name": product["sourceFileName"],
        "source_locators": json.dumps(product.get("sourceLocators") or [], separators=(",", ":")),
        "source_row_ids": json.dumps(product.get("sourceRowIds") or [], separators=(",", ":")),
        "ambiguity_codes": json.dumps(product.get("ambiguityCodes") or [], separators=(",", ":")),
        "review_notes": json.dumps(product.get("reviewNotes") or [], separators=(",", ":")),
        "bedrock_packet_id": bedrock.get("packetId"),
        "bedrock_prompt_version": bedrock.get("promptVersion"),
        "bedrock_request_hash": bedrock.get("requestHash"),
    }


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def slug(value: str) -> str:
    result = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return result or hashlib.sha256(value.encode()).hexdigest()[:12]


def require(path: Path, field: str) -> dict[str, Any]:
    value = read_json(path)
    if not value.get(field):
        raise RuntimeError(f"Required certificate {path.name} is invalid.")
    return value


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.resolve(strict=True).read_bytes()).hexdigest()


def read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.resolve(strict=True).read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"Expected object in {path}")
    return value


def write_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
